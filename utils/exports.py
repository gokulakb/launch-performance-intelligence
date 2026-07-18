"""
Export functionality for CSV, Excel, and PDF reports.
"""

import pandas as pd
import streamlit as st
import io
import base64
from typing import Dict, Any, List, Optional
from datetime import datetime
import logging
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

class ExportManager:
    """Manager for exporting data and reports."""
    
    @staticmethod
    def export_csv(df: pd.DataFrame, filename: str = "export") -> bytes:
        """
        Export DataFrame to CSV.
        
        Args:
            df: DataFrame to export
            filename: Output filename
        
        Returns:
            CSV bytes
        """
        if df.empty:
            return b""
        
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False)
        return csv_buffer.getvalue().encode('utf-8')
    
    @staticmethod
    def export_excel(df: pd.DataFrame, filename: str = "export", 
                     sheet_name: str = "Data") -> bytes:
        """
        Export DataFrame to Excel.
        
        Args:
            df: DataFrame to export
            filename: Output filename
            sheet_name: Excel sheet name
        
        Returns:
            Excel bytes
        """
        if df.empty:
            return b""
        
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Format Excel
            workbook = writer.book
            worksheet = workbook[sheet_name]
            
            # Add header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return excel_buffer.getvalue()
    
    @staticmethod
    def export_multiple_excel(dataframes: Dict[str, pd.DataFrame], 
                             filename: str = "export") -> bytes:
        """
        Export multiple DataFrames to Excel with multiple sheets.
        
        Args:
            dataframes: Dictionary of DataFrames (name -> DataFrame)
            filename: Output filename
        
        Returns:
            Excel bytes
        """
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            for sheet_name, df in dataframes.items():
                if not df.empty:
                    df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        
        return excel_buffer.getvalue()
    
    @staticmethod
    def export_pdf_report(title: str, sections: Dict[str, Any], 
                         filename: str = "report") -> bytes:
        """
        Export report to PDF.
        
        Args:
            title: Report title
            sections: Dictionary of sections (title -> content)
            filename: Output filename
        
        Returns:
            PDF bytes
        """
        pdf = FPDF()
        pdf.add_page()
        
        # Set up fonts
        pdf.set_font("Arial", "B", 16)
        
        # Add title
        pdf.cell(0, 10, title, 0, 1, "C")
        pdf.ln(10)
        
        # Add date
        pdf.set_font("Arial", "I", 10)
        pdf.cell(0, 10, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", 0, 1, "R")
        pdf.ln(5)
        
        # Add sections
        pdf.set_font("Arial", "B", 12)
        
        for section_title, content in sections.items():
            pdf.set_font("Arial", "B", 12)
            pdf.cell(0, 10, section_title, 0, 1)
            pdf.set_font("Arial", "", 10)
            
            # Handle different content types
            if isinstance(content, pd.DataFrame):
                # Convert DataFrame to table
                if not content.empty:
                    # Headers
                    headers = content.columns.tolist()
                    col_widths = [50] * len(headers)  # Adjust as needed
                    
                    # Print headers
                    for i, header in enumerate(headers):
                        pdf.cell(col_widths[i] if i < len(col_widths) else 30, 8, str(header), 1)
                    pdf.ln()
                    
                    # Print data rows (limit to 20 rows for PDF)
                    for _, row in content.head(20).iterrows():
                        for i, col in enumerate(headers):
                            value = str(row[col])[:20]  # Truncate long values
                            pdf.cell(col_widths[i] if i < len(col_widths) else 30, 6, value, 1)
                        pdf.ln()
                    
                    if len(content) > 20:
                        pdf.cell(0, 6, f"... and {len(content) - 20} more rows", 0, 1, "C")
                else:
                    pdf.cell(0, 6, "No data available", 0, 1)
            
            elif isinstance(content, dict):
                # Print dictionary as key-value pairs
                for key, value in content.items():
                    pdf.cell(60, 6, f"{key}:", 0)
                    pdf.cell(0, 6, str(value), 0, 1)
            
            elif isinstance(content, list):
                # Print list as bullet points
                for item in content[:20]:  # Limit items
                    pdf.cell(10, 6, "•", 0)
                    pdf.cell(0, 6, str(item)[:80], 0, 1)
                
                if len(content) > 20:
                    pdf.cell(0, 6, f"... and {len(content) - 20} more items", 0, 1)
            
            else:
                # Print as text
                pdf.multi_cell(0, 6, str(content)[:1000])
            
            pdf.ln(5)
        
        return pdf.output(dest='S').encode('latin1')
    
    @staticmethod
    def download_button(data: bytes, filename: str, mime_type: str, 
                       button_text: str = "Download") -> None:
        """
        Create a download button in Streamlit.
        
        Args:
            data: File data
            filename: Output filename
            mime_type: MIME type
            button_text: Button text
        """
        if data:
            st.download_button(
                label=button_text,
                data=data,
                file_name=filename,
                mime_type=mime_type
            )
    
    @staticmethod
    def export_summary_report(data: Dict[str, Any], filename: str = "summary_report") -> bytes:
        """
        Export a summary report to PDF.
        
        Args:
            data: Dictionary with report data
            filename: Output filename
        
        Returns:
            PDF bytes
        """
        sections = {
            "Executive Summary": data.get('summary', "No summary available"),
            "Key Metrics": data.get('metrics', {}),
            "Revenue Overview": data.get('revenue', pd.DataFrame()),
            "Retention Analysis": data.get('retention', pd.DataFrame()),
            "Data Quality": data.get('quality', {}),
            "Top Issues": data.get('issues', [])
        }
        
        return ExportManager.export_pdf_report(
            title="Launch Performance Summary Report",
            sections=sections,
            filename=filename
        )
    
    @staticmethod
    def get_download_link(data: bytes, filename: str, mime_type: str) -> str:
        """
        Generate an HTML download link.
        
        Args:
            data: File data
            filename: Output filename
            mime_type: MIME type
        
        Returns:
            HTML link
        """
        b64 = base64.b64encode(data).decode()
        href = f'<a href="data:{mime_type};base64,{b64}" download="{filename}">Download {filename}</a>'
        return href